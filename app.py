import openpyxl

wb = openpyxl.Workbook()
wb = openpyxl.load_workbook("Coding.xlsx")
print(wb.sheetnames)

sheet = wb["Sheet1"]

# wb.create_sheet("Sheet2", 0)
# wb.remove_sheet(sheet)
cell = sheet["a1"]
cell = sheet.cell(row=1, column=1)

print(sheet.max_row)
print(sheet.max_column)
for row in range(1, sheet.max_row + 1):
    for column in range(1, sheet.max_column + 1):
        cell = sheet.cell(row, column)

        print(cell.value)
